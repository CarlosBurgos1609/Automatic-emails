!pip install pandas openpyxl

from google.colab import drive
drive.mount('/content/drive')

import imaplib
import smtplib
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import time
import os
import pandas as pd
import openpyxl
from datetime import datetime

EMAIL_ACCOUNT = 'CORREOELCTRÓNICO'  # Reemplaza con tu correo
EMAIL_PASSWORD = 'CONTRASEÑADEAPLICACIÓN'  # Contraseña o contraseña de aplicación
IMAP_SERVER = 'imap-mail.outlook.com'
SMTP_SERVER = 'smtp-mail.outlook.com'

EXCEL_PATH = '/content/drive/MyDrive/Judicatura/Juzgados.xlsx'

# Diccionario para nombres de meses en español (mayúsculas)
MONTHS_ES = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL', 5: 'MAYO', 6: 'JUNIO',
    7: 'JULIO', 8: 'AGOSTO', 9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
}

def get_open_court():
    today = datetime.now()
    month_num = today.month
    month_es = MONTHS_ES.get(month_num, None)
    if not month_es:
        raise ValueError("Mes no válido")

    today_str = today.strftime('%d/%m/%Y')
    print(f"Buscando juzgado abierto en la hoja '{month_es}' para fecha: {today_str}")

    # Leer la hoja del mes
    df = pd.read_excel(EXCEL_PATH, sheet_name=month_es)
    print(f"Tabla de la hoja '{month_es}':\n{df.to_string(index=False)}")  # Mostrar tabla sin índice

    # Filtrar por fecha actual y estado "Abierto" (case insensitive)
    row = df[(df['FECHA'] == today_str) & (df['ESTADO'].str.lower() == 'abierto')]

    if not row.empty:
        court_name = row['JUZGADO'].iloc[0]
        court_email = row['CORREO'].iloc[0]
        municipio = row['Municipio'].iloc[0]
        print(f"Juzgado abierto encontrado: {court_name} ({court_email}, {municipio})")
        return court_name, court_email, municipio
    print("No se encontró juzgado abierto para hoy.")
    return None, None, None

def read_habeas_corpus_emails():
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
    mail.select('inbox')

    # Buscar correos no leídos con "habeas corpus" en el asunto
    _, message_numbers = mail.search(None, '(UNSEEN SUBJECT "habeas corpus")')
    messages = []

    for num in message_numbers[0].split():
        _, msg_data = mail.fetch(num, '(RFC822)')
        email_body = msg_data[0][1]
        msg = email.message_from_bytes(email_body)
        subject = msg['subject']
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == 'text/plain':
                    content = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                    break
        else:
            content = msg.get_payload(decode=True).decode('utf-8', errors='ignore')
        messages.append((num, subject, content))
        print(f"Correo encontrado: Asunto='{subject}', Contenido inicial={content[:50]}...")  # Mostrar primeros 50 chars

    return mail, messages

def send_email_to_court(to_email, subject, content):
    msg = MIMEMultipart()
    msg['From'] = EMAIL_ACCOUNT
    msg['To'] = to_email
    msg['Subject'] = f'Reenviado: {subject}'
    msg.attach(MIMEText(content, 'plain'))

    server = smtplib.SMTP(SMTP_SERVER, 587)
    server.starttls()
    server.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
    server.sendmail(EMAIL_ACCOUNT, to_email, msg.as_string())
    server.quit()
    print(f"Correo enviado a: {to_email}")

def mark_email_processed(mail, msg_id):
    mail.store(msg_id, '+FLAGS', '\Seen')
    print(f"Correo con ID {msg_id.decode('utf-8')} marcado como leído.")

def log_to_excel(subject, court_email, municipio):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['correos']  # Hoja "correos"

    # Encontrar la última fila
    last_row = ws.max_row
    new_row = last_row + 1

    # ID: Incrementar del anterior
    id_col = 1  # Columna A
    if last_row > 1:
        new_id = ws.cell(row=last_row, column=id_col).value + 1
    else:
        new_id = 1

    # Fecha y hora actual
    fecha = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    # Valores (ajusta "usuario" si necesitas algo diferente)
    usuario = "Sistema Automatico"
    correo_usuario = EMAIL_ACCOUNT
    archivo = ""  # Vacío si no hay archivo adjunto

    # Escribir en las columnas (A a H)
    ws.cell(row=new_row, column=1, value=new_id)  # id
    ws.cell(row=new_row, column=2, value=fecha)   # fecha y hora
    ws.cell(row=new_row, column=3).value = f'=TEXTO(B{new_row};"DDDD")'  # Fórmula para día
    ws.cell(row=new_row, column=4, value=usuario)  # usuario
    ws.cell(row=new_row, column=5, value=correo_usuario)  # correo usuario
    ws.cell(row=new_row, column=6, value=court_email)  # correo juzgado
    ws.cell(row=new_row, column=7, value=archivo)  # archivo
    ws.cell(row=new_row, column=8, value=municipio)  # municipio

    # Guardar el archivo
    wb.save(EXCEL_PATH)
    print(f"Registro agregado en fila {new_row}: ID={new_id}, Fecha={fecha}, Juzgado={court_email}")

def main():
    while True:
        try:
            # Obtener juzgado abierto
            court_name, court_email, municipio = get_open_court()
            if not court_email:
                print("No hay juzgados abiertos hoy.")
                time.sleep(300)
                continue

            # Leer correos
            mail, messages = read_habeas_corpus_emails()
            for msg_id, subject, content in messages:
                if "habeas corpus" in subject.lower():
                    print(f"Procesando correo: {subject}")
                    send_email_to_court(court_email, subject, content)
                    mark_email_processed(mail, msg_id)
                    log_to_excel(subject, court_email, municipio)
                    print(f"Correo enviado a {court_name} ({court_email}) y registrado en Excel.")

            mail.logout()
            print("Ciclo completado. Esperando 5 minutos...")
            time.sleep(300)

        except Exception as error:
            print(f"Error: {error}")
            time.sleep(60)  # Reintentar en 1 minuto

if __name__ == '__main__':
    main()
